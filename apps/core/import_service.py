"""
import_service.py — Version web (Django / PostgreSQL)
Même logique que le desktop mais utilise les modèles Django.

Supporte deux formats :
  1. Format original Report_Carburant.xlsx
     - Feuille "Rapport journalier" + feuilles "Synthèse"
  2. Format export CarbPro Web
     - Feuille "Synthèse" + feuille "Rapport journalier"
"""

import re
from datetime import datetime

# ── Mapping type Excel → (type_interne, mode_appro) ──────────
TYPE_MAP = {
    "CAMION":          ("camion_benne",    "avec_index"),
    "CAMION BENNE":    ("camion_benne",    "avec_index"),
    "CHARGEUR":        ("chargeur",        "avec_index"),
    "EXCAVATRICE":     ("excavatrice",     "avec_index"),
    "NEW BULLDOZER":   ("bulldozer",       "avec_index"),
    "BULLDOZER":       ("bulldozer",       "avec_index"),
    "NIVELEUSE":       ("niveleuse",       "avec_index"),
    "COMPACTEUR":      ("compacteur",      "avec_index"),
    "WATER TANK":      ("equipement_fixe", "sans_index"),
    "CAMION CITERNE":  ("equipement_fixe", "sans_index"),
    "PORTE-CHAR":      ("autre",           "sans_index"),
    "LAND CR. DIR":    ("vehicule",        "sans_index"),
}

TYPES_DIVERSES = {
    "FUEL TANK", "BIDON", "BIDON GARAGE", "GARAGE",
    "CAMION GARAGE", "TLB", "GROUPE ELEC", "AUTRES",
}

CATEGORIE_MAP = {
    "FUEL TANK":      "Fuel Tank",
    "BIDON":          "Bidon",
    "BIDON GARAGE":   "Garage",
    "GARAGE":         "Garage",
    "CAMION GARAGE":  "Garage",
    "TLB":            "Autre",
    "GROUPE ELEC":    "Groupe électrogène",
    "AUTRES":         "Autre",
}

NORMES = {
    "camion_benne": ("km", 2.0, 1.9),
    "excavatrice":  ("h",  25.0, None),
    "chargeur":     ("h",  20.0, None),
    "bulldozer":    ("h",  27.0, None),
    "niveleuse":    ("h",  14.0, None),
    "compacteur":   ("h",  12.0, None),
}


def _parse_index(val):
    if val is None:
        return None, False
    if isinstance(val, (int, float)):
        return float(val), False
    s = str(val).strip()
    if not s:
        return None, False
    if "panne" in s.lower():
        return None, True
    if s.startswith("="):
        return None, False
    s_clean = re.sub(r'[kKmMhH\s]', '', s).replace(',', '.')
    try:
        return float(s_clean), False
    except ValueError:
        return None, False


def _parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if hasattr(val, 'date'):
        return val.date()
    try:
        from datetime import date
        return datetime.strptime(str(val)[:10], "%Y-%m-%d").date()
    except Exception:
        try:
            return datetime.strptime(str(val)[:10], "%d/%m/%Y").date()
        except Exception:
            return None


def analyser_fichier(filepath):
    """
    Analyse le fichier Excel et retourne les données structurées.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)

    result = {
        "entrees_stock":          [],
        "rav_engins":             [],
        "consommations_diverses": [],
        "erreurs":                [],
    }

    # ── 1. Entrées stock (feuilles contenant "synth") ─────────
    for sheet_name in wb.sheetnames:
        if "synth" in sheet_name.lower():
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                date_val = None
                qte_val  = None
                for i, cell in enumerate(row):
                    if date_val is None and (
                        isinstance(cell, datetime) or
                        (isinstance(cell, str) and len(str(cell)) >= 8)
                    ):
                        d = _parse_date(cell)
                        if d:
                            date_val = d
                    if i == 2 and isinstance(cell, (int, float)) and cell > 0:
                        qte_val = cell
                if date_val and qte_val:
                    result["entrees_stock"].append({
                        "date":     date_val,
                        "quantite": float(qte_val),
                        "source":   sheet_name,
                    })

    # Dédupliquer entrées stock par date
    entrees_map = {}
    for e in result["entrees_stock"]:
        d = e["date"]
        if d not in entrees_map or e["quantite"] > entrees_map[d]["quantite"]:
            entrees_map[d] = e
    result["entrees_stock"] = sorted(entrees_map.values(), key=lambda x: x["date"])

    # ── 2. Rapport journalier ──────────────────────────────────
    sheet_rj = None
    for name in wb.sheetnames:
        if "journalier" in name.lower() or "rapport" in name.lower():
            sheet_rj = wb[name]
            break

    if sheet_rj is None:
        result["erreurs"].append(
            "Feuille 'Rapport journalier' introuvable — "
            "seules les entrées stock ont été importées."
        )
        return result

    last_date = None

    for row in sheet_rj.iter_rows(min_row=12, values_only=True):
        # Ignorer les lignes vides ou sous-totaux
        if not any(row):
            continue
        qte_cell = row[8] if len(row) > 8 else None
        if qte_cell is None:
            continue
        if isinstance(qte_cell, str) and (
            qte_cell.startswith("=") or qte_cell.startswith("+")
        ):
            continue

        # Date
        date_val = _parse_date(row[1]) if len(row) > 1 else None
        if date_val:
            last_date = date_val
        else:
            date_val = last_date
        if not date_val:
            continue

        # Type et ID engin
        type_excel = str(row[4]).strip().upper() if len(row) > 4 and row[4] else None
        id_engin   = str(row[5]).strip().upper() if len(row) > 5 and row[5] else None
        idx_prec   = row[6] if len(row) > 6 else None
        idx_act    = row[7] if len(row) > 7 else None
        obs        = str(row[9]).strip() if len(row) > 9 and row[9] else ""

        if not type_excel:
            continue

        # Quantité
        try:
            qte = float(qte_cell)
        except (ValueError, TypeError):
            continue
        if qte <= 0:
            continue

        # Nettoyer id_engin
        if id_engin in ("NONE", "NEW", "MG", "", "NAN"):
            id_engin = None

        # ── Consommation diverse ──────────────────────────────
        if type_excel in TYPES_DIVERSES or id_engin is None:
            cat = CATEGORIE_MAP.get(type_excel, "Autre")
            result["consommations_diverses"].append({
                "date":      date_val,
                "categorie": cat,
                "qte":       qte,
                "motif":     obs or type_excel,
            })
            continue

        # ── Engin identifié ───────────────────────────────────
        mapping = TYPE_MAP.get(type_excel) or TYPE_MAP.get(type_excel.strip())
        if not mapping:
            cat = CATEGORIE_MAP.get(type_excel, "Autre")
            result["consommations_diverses"].append({
                "date":      date_val,
                "categorie": cat,
                "qte":       qte,
                "motif":     f"{type_excel} {id_engin or ''}".strip(),
            })
            continue

        type_interne, mode_appro = mapping
        idx_prec_val, panne1 = _parse_index(idx_prec)
        idx_act_val,  panne2 = _parse_index(idx_act)
        panne_index = panne1 or panne2

        if mode_appro == "sans_index":
            panne_index  = False
            idx_prec_val = None
            idx_act_val  = None

        premier_plein = (
            mode_appro == "avec_index"
            and not panne_index
            and idx_prec_val is None
            and idx_act_val is not None
        )

        result["rav_engins"].append({
            "date":          date_val,
            "id_engin":      id_engin,
            "type_excel":    type_excel,
            "type_interne":  type_interne,
            "mode_appro":    mode_appro,
            "index_prec":    idx_prec_val,
            "index_act":     idx_act_val,
            "qte":           qte,
            "obs":           obs,
            "panne_index":   panne_index,
            "premier_plein": premier_plein,
        })

    return result


def importer_depuis_excel(filepath, operateur=None):
    """
    Importe les données Excel via bulk_create — optimisé pour éviter les timeouts.
    """
    from .models import Engin, OperationStock, RavitaillementEngin, ConsommationDiverse

    analyse = analyser_fichier(filepath)
    rapport = {
        "entrees_stock_ok":   0,
        "entrees_stock_skip": 0,
        "rav_ok":             0,
        "rav_skip":           0,
        "diverses_ok":        0,
        "engins_crees":       [],
        "erreurs":            list(analyse["erreurs"]),
    }

    # ── 1. Entrées stock — bulk_create ────────────────────────
    dates_existantes = set(
        OperationStock.objects.filter(type="entree").values_list("date", flat=True)
    )
    nouvelles_entrees = []
    for e in analyse["entrees_stock"]:
        if e["date"] in dates_existantes:
            rapport["entrees_stock_skip"] += 1
            continue
        nouvelles_entrees.append(OperationStock(
            date        = e["date"],
            type        = "entree",
            quantite    = e["quantite"],
            stock_apres = 0,
            description = f"Import — {e['source']}",
            operateur   = operateur,
        ))
    if nouvelles_entrees:
        OperationStock.objects.bulk_create(nouvelles_entrees, ignore_conflicts=True)
        rapport["entrees_stock_ok"] = len(nouvelles_entrees)

    # ── 2. Créer les engins manquants en une passe ────────────
    ids_existants = set(Engin.objects.values_list("id_engin", flat=True))
    engins_a_creer = {}
    for rv in analyse["rav_engins"]:
        if rv["id_engin"] not in ids_existants and rv["id_engin"] not in engins_a_creer:
            engins_a_creer[rv["id_engin"]] = Engin(
                id_engin    = rv["id_engin"],
                type_engin  = rv["type_interne"],
                description = f"Importé — {rv['type_excel']}",
                mode_appro  = rv["mode_appro"],
                actif       = True,
            )
    if engins_a_creer:
        Engin.objects.bulk_create(list(engins_a_creer.values()), ignore_conflicts=True)
        rapport["engins_crees"] = list(engins_a_creer.keys())

    # ── 3. Ravitaillements engins — bulk_create ────────────────
    # Charger tous les engins en mémoire (dict id_engin → objet)
    engins_map = {e.id_engin: e for e in Engin.objects.all()}

    ravs_a_creer   = []
    sorties_a_creer = []

    for rv in analyse["rav_engins"]:
        try:
            engin = engins_map.get(rv["id_engin"])
            if not engin:
                continue

            if rv["panne_index"]:
                statut = "panne_index"; idx_prec = idx_act = 0
                taux_reel = norme_ref = None
                commentaire = f"[PANNE INDEX] {rv['obs']}".strip()
            elif rv["mode_appro"] == "sans_index":
                statut = "non_verifie"; idx_prec = idx_act = 0
                taux_reel = norme_ref = None
                commentaire = rv["obs"]
            else:
                idx_prec = rv["index_prec"] or 0
                idx_act  = rv["index_act"]  or idx_prec
                diff     = idx_act - idx_prec
                taux_reel = norme_ref = None
                statut    = "non_verifie"
                commentaire = rv["obs"]
                info = NORMES.get(rv["type_interne"])
                if info and diff > 0:
                    unite, norme, seuil_min = info
                    norme_ref = norme
                    if unite == "km":
                        taux_reel = diff / rv["qte"] if rv["qte"] else 0
                        statut = "anomalie" if seuil_min and taux_reel < seuil_min else "normal"
                    else:
                        taux_reel = rv["qte"] / diff if diff else 0
                        statut = "anomalie" if taux_reel > norme * 1.1 else "normal"
                if rv["obs"] and "anormal" in rv["obs"].lower():
                    statut = "anomalie"

            ravs_a_creer.append(RavitaillementEngin(
                date             = rv["date"],
                engin            = engin,
                index_precedent  = idx_prec,
                index_actuel     = idx_act,
                difference_index = idx_act - idx_prec,
                qte_donnee       = rv["qte"],
                taux_reel        = taux_reel,
                norme_ref        = norme_ref,
                statut           = statut,
                commentaire      = commentaire or "",
                operateur        = operateur,
            ))
            sorties_a_creer.append(OperationStock(
                date        = rv["date"],
                type        = "sortie",
                quantite    = rv["qte"],
                stock_apres = 0,
                description = f"Appro {rv['id_engin']} — import",
                operateur   = operateur,
            ))
        except Exception as ex:
            rapport["erreurs"].append(f"Engin {rv['id_engin']} {rv['date']} : {ex}")
            rapport["rav_skip"] += 1

    if ravs_a_creer:
        RavitaillementEngin.objects.bulk_create(ravs_a_creer, ignore_conflicts=True)
        rapport["rav_ok"] = len(ravs_a_creer)

    # ── 4. Consommations diverses — bulk_create ────────────────
    div_a_creer     = []
    sorties_div     = []
    for d in analyse["consommations_diverses"]:
        div_a_creer.append(ConsommationDiverse(
            date        = d["date"],
            categorie   = d["categorie"],
            quantite    = d["qte"],
            motif       = f"Import — {d['motif']}" if d["motif"] else "Import",
            operateur   = operateur,
            stock_apres = 0,
        ))
        sorties_div.append(OperationStock(
            date        = d["date"],
            type        = "sortie",
            quantite    = d["qte"],
            stock_apres = 0,
            description = f"Divers import — {d['categorie']}",
            operateur   = operateur,
        ))
    if div_a_creer:
        ConsommationDiverse.objects.bulk_create(div_a_creer, ignore_conflicts=True)
        rapport["diverses_ok"] = len(div_a_creer)

    # ── 5. Toutes les sorties en une fois ──────────────────────
    all_sorties = sorties_a_creer + sorties_div
    if all_sorties:
        OperationStock.objects.bulk_create(all_sorties, ignore_conflicts=True)

    # ── 6. Recalcul stock SQL (1 requête) ──────────────────────
    _recalculer_stocks()
    return rapport


def _recalculer_stocks():
    """
    Recalcule stock_apres via SQL pur — sans charger les objets en mémoire.
    Utilise une window function PostgreSQL pour la somme cumulée.
    """
    from django.db import connection
    with connection.cursor() as cursor:
        cursor.execute("""
            UPDATE core_operationstock AS op
            SET stock_apres = sub.cumul
            FROM (
                SELECT
                    id,
                    SUM(
                        CASE WHEN type = 'entree' THEN quantite ELSE -quantite END
                    ) OVER (ORDER BY date, cree_le ROWS UNBOUNDED PRECEDING) AS cumul
                FROM core_operationstock
            ) AS sub
            WHERE op.id = sub.id
        """)
