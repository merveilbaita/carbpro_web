from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import HttpResponse, FileResponse
from django.utils import timezone
from django.db.models import Sum
from django.views.decorators.cache import cache_control
from django.views.decorators.http import require_GET
from functools import wraps
import json, openpyxl, os
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

from .models import (
    OperationStock, RavitaillementEngin,
    ConsommationDiverse, Engin, Parametre, UserProfile
)
from .forms import (
    RavitaillementStockForm, ApproEnginForm, ConsommationDiverseForm,
    CreerUtilisateurForm, ModifierUtilisateurForm,
)


# ── Décorateur rôle admin ─────────────────────────────────────
def admin_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        try:
            role = request.user.profile.role
        except Exception:
            role = "operateur"
        if role != "administrateur" and not request.user.is_superuser:
            messages.error(request, "⛔ Accès réservé aux administrateurs.")
            return redirect("dashboard")
        return view_func(request, *args, **kwargs)
    return login_required(wrapper)


def get_user_role(user):
    try:
        return user.profile.role
    except Exception:
        return "operateur"


# ── PWA : Service Worker servi depuis la racine ───────────────
@require_GET
@cache_control(max_age=0, no_cache=True, no_store=True, must_revalidate=True)
def service_worker(request):
    sw_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        '..', 'static', 'sw.js'
    )
    # Chercher dans staticfiles (production) ou static (dev)
    for candidate in [
        os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(
            os.path.abspath(__file__)))), 'staticfiles', 'sw.js'),
        os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(
            os.path.abspath(__file__)))), 'static', 'sw.js'),
    ]:
        if os.path.exists(candidate):
            with open(candidate, 'r') as f:
                return HttpResponse(f.read(),
                    content_type='application/javascript')
    return HttpResponse('// SW not found', content_type='application/javascript')


# ── PWA : Page offline ────────────────────────────────────────
def offline(request):
    return render(request, 'core/offline.html')

from .models import (
    OperationStock, RavitaillementEngin,
    ConsommationDiverse, Engin, Parametre, UserProfile
)
from .forms import (
    RavitaillementStockForm, ApproEnginForm, ConsommationDiverseForm,
    CreerUtilisateurForm, ModifierUtilisateurForm,
)


# ── Helpers ───────────────────────────────────────────────────

def get_stock_actuel():
    """Calcule le stock actuel depuis toutes les opérations."""
    entrees = OperationStock.objects.filter(type="entree").aggregate(
        t=Sum("quantite"))["t"] or 0
    sorties_stock = OperationStock.objects.filter(type="sortie").aggregate(
        t=Sum("quantite"))["t"] or 0
    sorties_engin = RavitaillementEngin.objects.aggregate(
        t=Sum("qte_donnee"))["t"] or 0
    sorties_div = ConsommationDiverse.objects.aggregate(
        t=Sum("quantite"))["t"] or 0
    return entrees - sorties_stock - sorties_engin - sorties_div


def recalc_stock_apres(op_stock):
    """Recalcule stock_apres pour une OperationStock."""
    ops = OperationStock.objects.order_by("date", "cree_le")
    stock = 0
    for op in ops:
        if op.type == "entree":
            stock += op.quantite
        else:
            stock -= op.quantite
        op.stock_apres = stock
        op.save(update_fields=["stock_apres"])


# ── Dashboard ─────────────────────────────────────────────────

@login_required
def dashboard(request):
    stock   = get_stock_actuel()
    now     = timezone.localdate()
    mois    = now.month
    annee   = now.year

    entrees_mois = OperationStock.objects.filter(
        type="entree", date__month=mois, date__year=annee
    ).aggregate(t=Sum("quantite"))["t"] or 0

    sorties_engin_mois = RavitaillementEngin.objects.filter(
        date__month=mois, date__year=annee
    ).aggregate(t=Sum("qte_donnee"))["t"] or 0

    sorties_div_mois = ConsommationDiverse.objects.filter(
        date__month=mois, date__year=annee
    ).aggregate(t=Sum("quantite"))["t"] or 0

    sorties_mois = sorties_engin_mois + sorties_div_mois

    nb_engins    = Engin.objects.filter(actif=True).count()
    dernieres_ops = OperationStock.objects.select_related("operateur")[:5]
    derniers_ravs = RavitaillementEngin.objects.select_related(
        "engin", "operateur")[:5]

    ctx = {
        "stock":          stock,
        "stock_negatif":  stock < 0,
        "stock_bas":      0 < stock <= float(Parametre.get("seuil_alerte_stock", "500")),
        "seuil":          float(Parametre.get("seuil_alerte_stock", "500")),
        "entrees_mois":   entrees_mois,
        "sorties_mois":   sorties_mois,
        "nb_engins":      nb_engins,
        "dernieres_ops":  dernieres_ops,
        "derniers_ravs":  derniers_ravs,
        "mois_label":     now.strftime("%B %Y"),
    }
    return render(request, "core/dashboard.html", ctx)


# ── Ravitaillement Stock ───────────────────────────────────────

@login_required
def ravitaillement_stock(request):
    if request.method == "POST":
        form = RavitaillementStockForm(request.POST)
        if form.is_valid():
            op = form.save(commit=False)
            op.type      = "entree"
            op.operateur = request.user
            # Calcul stock_apres
            stock_avant = get_stock_actuel()
            op.stock_apres = stock_avant + op.quantite
            op.save()
            messages.success(request,
                f"✅ Entrée de {op.quantite:,.0f} L enregistrée. "
                f"Stock actuel : {op.stock_apres:,.0f} L")
            return redirect("dashboard")
    else:
        form = RavitaillementStockForm()

    return render(request, "core/ravitaillement_stock.html", {
        "form":  form,
        "stock": get_stock_actuel(),
    })


# ── Approvisionnement Engin ────────────────────────────────────

@login_required
def appro_engin(request):
    if request.method == "POST":
        form = ApproEnginForm(request.POST)
        if form.is_valid():
            rav          = form.save(commit=False)
            rav.operateur = request.user
            engin         = rav.engin
            panne         = form.cleaned_data.get("panne_index")
            premier       = form.cleaned_data.get("premier_plein")

            # Statut
            if panne:
                rav.statut = "panne_index"
            elif engin.mode_appro == "sans_index" or premier:
                rav.statut = "non_verifie"
            else:
                diff = (rav.index_actuel or 0) - (rav.index_precedent or 0)
                rav.difference_index = diff
                # Calcul taux réel selon type
                NORMES = {
                    "camion_benne": ("km", 2.0, 1.9),
                    "excavatrice":  ("h",  25.0, None),
                    "chargeur":     ("h",  20.0, None),
                    "bulldozer":    ("h",  27.0, None),
                    "niveleuse":    ("h",  14.0, None),
                    "compacteur":   ("h",  12.0, None),
                }
                info = NORMES.get(engin.type_engin)
                if info and diff > 0:
                    unite, norme, seuil_min = info
                    if unite == "km":
                        taux = diff / rav.qte_donnee if rav.qte_donnee else 0
                        rav.taux_reel = round(taux, 3)
                        rav.norme_ref = norme
                        rav.statut    = "anomalie" if seuil_min and taux < seuil_min else "normal"
                    else:
                        taux = rav.qte_donnee / diff if diff else 0
                        rav.taux_reel = round(taux, 3)
                        rav.norme_ref = norme
                        rav.statut    = "anomalie" if taux > norme * 1.1 else "normal"
                else:
                    rav.statut = "non_verifie"

            # Sortie stock
            stock_avant = get_stock_actuel()
            stock_apres = stock_avant - rav.qte_donnee
            rav.save()

            # Enregistrer la sortie correspondante
            OperationStock.objects.create(
                date        = rav.date,
                type        = "sortie",
                quantite    = rav.qte_donnee,
                stock_apres = stock_apres,
                description = f"Appro {engin.id_engin}",
                operateur   = request.user,
            )

            tag_stock = f"⚠️ Stock négatif : {stock_apres:,.0f} L" if stock_apres < 0 \
                        else f"Stock restant : {stock_apres:,.0f} L"
            messages.success(request,
                f"✅ Ravitaillement {engin.id_engin} enregistré ({rav.qte_donnee:,.0f} L). "
                f"{tag_stock}")
            return redirect("dashboard")
    else:
        form = ApproEnginForm()

    # Données engins pour le JS (mode_appro + dernier index)
    engins_data = {}
    for e in Engin.objects.filter(actif=True):
        dernier = RavitaillementEngin.objects.filter(
            engin=e).order_by("-date", "-cree_le").first()
        engins_data[e.id_engin] = {
            "mode":          e.mode_appro,
            "type":          e.type_engin,
            "dernier_index": dernier.index_actuel if dernier else 0,
        }

    return render(request, "core/appro_engin.html", {
        "form":        form,
        "stock":       get_stock_actuel(),
        "engins_data": json.dumps(engins_data),
    })


# ── Consommation Diverse ──────────────────────────────────────

@login_required
def consommation_diverse(request):
    if request.method == "POST":
        form = ConsommationDiverseForm(request.POST)
        if form.is_valid():
            div           = form.save(commit=False)
            div.operateur = request.user
            stock_avant   = get_stock_actuel()
            div.stock_apres = stock_avant - div.quantite
            div.save()

            OperationStock.objects.create(
                date        = div.date,
                type        = "sortie",
                quantite    = div.quantite,
                stock_apres = div.stock_apres,
                description = f"Divers — {div.categorie}",
                operateur   = request.user,
            )

            messages.success(request,
                f"✅ Consommation {div.categorie} enregistrée ({div.quantite:,.0f} L).")
            return redirect("dashboard")
    else:
        form = ConsommationDiverseForm()

    return render(request, "core/consommation_diverse.html", {
        "form":  form,
        "stock": get_stock_actuel(),
    })


# ── Historique ────────────────────────────────────────────────

@login_required
def historique(request):
    mois  = request.GET.get("mois",  timezone.localdate().month)
    annee = request.GET.get("annee", timezone.localdate().year)

    ops   = OperationStock.objects.filter(
        date__month=mois, date__year=annee
    ).select_related("operateur").order_by("-date")

    ravs  = RavitaillementEngin.objects.filter(
        date__month=mois, date__year=annee
    ).select_related("engin", "operateur").order_by("-date")

    divs  = ConsommationDiverse.objects.filter(
        date__month=mois, date__year=annee
    ).select_related("operateur").order_by("-date")

    annees = range(2024, timezone.localdate().year + 2)

    ctx = {
        "ops": ops, "ravs": ravs, "divs": divs,
        "mois": int(mois), "annee": int(annee),
        "annees": annees,
        "is_admin": request.user.is_superuser or get_user_role(request.user) == "administrateur",
        "mois_choices": [
            (1,"Janvier"),(2,"Février"),(3,"Mars"),(4,"Avril"),
            (5,"Mai"),(6,"Juin"),(7,"Juillet"),(8,"Août"),
            (9,"Septembre"),(10,"Octobre"),(11,"Novembre"),(12,"Décembre"),
        ],
    }
    return render(request, "core/historique.html", ctx)


# ── Export Excel (compatible import desktop) ──────────────────

@login_required
def export_excel(request):
    mois  = request.GET.get("mois",  timezone.localdate().month)
    annee = request.GET.get("annee", timezone.localdate().year)

    wb = openpyxl.Workbook()

    BLEU   = "1A3A6B"
    JAUNE  = "FFC000"

    # ── Feuille 1 : Synthèse entrées (lue par l'import desktop) ──
    # L'import desktop cherche les feuilles contenant "synth"
    # et lit : col 0=?, col 1=date, col 2=quantite
    ws_s = wb.active
    ws_s.title = "Synthèse"

    ws_s.append(["", "Date", "Quantité (L)", "Description", "Opérateur"])
    for cell in ws_s[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLEU)
        cell.alignment = Alignment(horizontal="center")

    for op in OperationStock.objects.filter(
        type="entree", date__month=mois, date__year=annee
    ).order_by("date"):
        ws_s.append([
            "",
            op.date,          # col index 1 — datetime lu par l'import
            op.quantite,      # col index 2 — quantite
            op.description,
            op.operateur.username if op.operateur else "",
        ])
        # Formater la cellule date comme datetime pour que openpyxl la reconnaisse
        ws_s.cell(row=ws_s.max_row, column=2).number_format = "DD/MM/YYYY"

    # ── Feuille 2 : Rapport journalier (lue par l'import desktop) ─
    # Structure attendue (min_row=12, index base 0) :
    # col 1 = date, col 4 = type_excel, col 5 = id_engin,
    # col 6 = index_prec, col 7 = index_act, col 8 = qte, col 9 = obs
    ws_r = wb.create_sheet("Rapport journalier")

    # Lignes 1-11 : en-tête (l'import commence à row 12)
    ws_r.append(["RAPPORT JOURNALIER — Export CarbPro Web"])
    for _ in range(10):
        ws_r.append([])  # lignes vides jusqu'à row 11

    # Ligne 12 : en-tête colonnes
    header = ["", "Date", "", "", "Type", "ID Engin",
              "Index Préc.", "Index Act.", "Quantité (L)", "Observations"]
    ws_r.append(header)
    for cell in ws_r[12]:
        if cell.value:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=BLEU)
            cell.alignment = Alignment(horizontal="center")

    # Type mapping inverse : type_interne → type_excel pour l'import
    TYPE_EXCEL_MAP = {
        "camion_benne":    "CAMION BENNE",
        "excavatrice":     "EXCAVATRICE",
        "chargeur":        "CHARGEUR",
        "bulldozer":       "BULLDOZER",
        "niveleuse":       "NIVELEUSE",
        "compacteur":      "COMPACTEUR",
        "vehicule":        "LAND CR. DIR",
        "equipement_fixe": "WATER TANK",
        "autre":           "PORTE-CHAR",
    }

    # Ravitaillements engins
    for rav in RavitaillementEngin.objects.filter(
        date__month=mois, date__year=annee
    ).select_related("engin").order_by("date"):
        type_excel = TYPE_EXCEL_MAP.get(rav.engin.type_engin, rav.engin.type_engin.upper())
        idx_prec   = rav.index_precedent if rav.engin.mode_appro == "avec_index" else ""
        idx_act    = rav.index_actuel    if rav.engin.mode_appro == "avec_index" else ""
        if rav.statut == "panne_index":
            idx_prec = "Panne"
            idx_act  = "Panne"
        row = ["", rav.date, "", "", type_excel, rav.engin.id_engin,
               idx_prec, idx_act, rav.qte_donnee, rav.commentaire or ""]
        ws_r.append(row)
        ws_r.cell(row=ws_r.max_row, column=2).number_format = "DD/MM/YYYY"

    # Consommations diverses — mappées vers les types diverses attendus
    CAT_EXCEL_MAP = {
        "Garage":                 "GARAGE",
        "Groupe électrogène":     "GROUPE ELEC",
        "Bidon":                  "BIDON",
        "Fuel Tank":              "FUEL TANK",
        "Land Cruiser Direction": "LAND CR. DIR",
        "Autre":                  "AUTRES",
    }
    for div in ConsommationDiverse.objects.filter(
        date__month=mois, date__year=annee
    ).order_by("date"):
        type_excel = CAT_EXCEL_MAP.get(div.categorie, "AUTRES")
        row = ["", div.date, "", "", type_excel, "",
               "", "", div.quantite, div.motif or ""]
        ws_r.append(row)
        ws_r.cell(row=ws_r.max_row, column=2).number_format = "DD/MM/YYYY"

    # ── Feuille 3 : Résumé lisible (pour consultation) ───────────
    ws_info = wb.create_sheet("Résumé Web")
    ws_info.append(["Export CarbPro Web",
                    f"Période : {mois}/{annee}",
                    f"Généré le {date.today().strftime('%d/%m/%Y')}"])
    ws_info.append([])
    ws_info.append(["Ce fichier est compatible avec l'import de GestionCarburantPro Desktop"])

    # Ajuster largeurs
    for ws in [ws_s, ws_r]:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="carbpro_export_{annee}_{int(mois):02d}.xlsx"'
    )
    wb.save(response)
    return response


# ── Engins (lecture seule pour les opérateurs terrain) ────────

@login_required
def liste_engins(request):
    engins = Engin.objects.filter(actif=True).order_by("id_engin")
    return render(request, "core/liste_engins.html", {"engins": engins})

@login_required
def import_excel(request):
    rapport = None

    if request.method == "POST" and request.FILES.get("fichier"):
        import tempfile, os
        fichier = request.FILES["fichier"]

        # Vérifier extension
        if not fichier.name.endswith((".xlsx", ".xls")):
            messages.error(request, "❌ Format invalide. Seuls les fichiers .xlsx sont acceptés.")
            return redirect("import_excel")

        # Sauvegarder temporairement
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            for chunk in fichier.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        try:
            from .import_service import importer_depuis_excel
            rapport = importer_depuis_excel(tmp_path, operateur=request.user)

            total_ok = (rapport["entrees_stock_ok"] +
                        rapport["rav_ok"] +
                        rapport["diverses_ok"])

            if total_ok > 0:
                messages.success(request,
                    f"✅ Import terminé : {rapport['entrees_stock_ok']} entrées stock, "
                    f"{rapport['rav_ok']} ravitaillements engins, "
                    f"{rapport['diverses_ok']} consommations diverses."
                )
            else:
                messages.warning(request,
                    "⚠️ Aucune donnée importée. Vérifiez le format du fichier.")

        except Exception as ex:
            messages.error(request, f"❌ Erreur lors de l'import : {ex}")
        finally:
            os.unlink(tmp_path)

    return render(request, "core/import_excel.html", {"rapport": rapport})
# ── Import Excel ───────────────────────────────────────────────

@login_required
def import_excel(request):
    rapport = None

    if request.method == "POST" and request.FILES.get("fichier"):
        import tempfile, os
        fichier = request.FILES["fichier"]

        if not fichier.name.endswith((".xlsx", ".xls")):
            messages.error(request, "❌ Format invalide. Seuls les fichiers .xlsx sont acceptés.")
            return redirect("import_excel")

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            for chunk in fichier.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        try:
            from .import_service import importer_depuis_excel
            rapport = importer_depuis_excel(tmp_path, operateur=request.user)

            total_ok = (rapport["entrees_stock_ok"] +
                        rapport["rav_ok"] +
                        rapport["diverses_ok"])

            if total_ok > 0:
                messages.success(request,
                    f"✅ Import terminé : {rapport['entrees_stock_ok']} entrées stock, "
                    f"{rapport['rav_ok']} ravitaillements engins, "
                    f"{rapport['diverses_ok']} consommations diverses."
                )
            else:
                messages.warning(request,
                    "⚠️ Aucune donnée importée. Vérifiez le format du fichier.")

        except Exception as ex:
            messages.error(request, f"❌ Erreur lors de l'import : {ex}")
        finally:
            os.unlink(tmp_path)

    return render(request, "core/import_excel.html", {"rapport": rapport})


# ── Gestion Utilisateurs ──────────────────────────────────────

@admin_required
def liste_utilisateurs(request):
    users = User.objects.select_related("profile").order_by("username")
    ctx = {
        "users":       users,
        "total":       users.count(),
        "nb_admins":   users.filter(profile__role="administrateur").count(),
        "nb_operateurs": users.filter(profile__role="operateur").count(),
    }
    return render(request, "core/utilisateurs/liste.html", ctx)


@admin_required
def creer_utilisateur(request):
    if request.method == "POST":
        form = CreerUtilisateurForm(request.POST)
        if form.is_valid():
            d = form.cleaned_data
            # Séparer nom complet en first/last name
            parts = d["nom_complet"].strip().split(" ", 1)
            first = parts[0]
            last  = parts[1] if len(parts) > 1 else ""

            user = User.objects.create_user(
                username   = d["username"],
                password   = d["password1"],
                email      = d.get("email", ""),
                first_name = first,
                last_name  = last,
            )
            # Profil rôle
            profile, _ = UserProfile.objects.get_or_create(user=user)
            profile.role = d["role"]
            profile.save()

            # Si admin → is_staff aussi
            if d["role"] == "administrateur":
                user.is_staff = True
                user.save()

            messages.success(request,
                f"✅ Utilisateur '{user.username}' créé avec le rôle {profile.get_role_display()}.")
            return redirect("liste_utilisateurs")
    else:
        form = CreerUtilisateurForm()

    return render(request, "core/utilisateurs/form.html", {
        "form":  form,
        "titre": "Créer un utilisateur",
        "mode":  "creation",
    })


@admin_required
def modifier_utilisateur(request, user_id):
    target = get_object_or_404(User, pk=user_id)
    profile, _ = UserProfile.objects.get_or_create(user=target)

    # Empêcher de modifier le superuser si on n'est pas superuser
    if target.is_superuser and not request.user.is_superuser:
        messages.error(request, "⛔ Vous ne pouvez pas modifier le superadmin.")
        return redirect("liste_utilisateurs")

    if request.method == "POST":
        form = ModifierUtilisateurForm(request.POST)
        if form.is_valid():
            d = form.cleaned_data
            parts = d["nom_complet"].strip().split(" ", 1)
            target.first_name = parts[0]
            target.last_name  = parts[1] if len(parts) > 1 else ""
            target.email      = d.get("email", "")

            if d.get("password_nouveau"):
                target.set_password(d["password_nouveau"])

            # Rôle
            profile.role = d["role"]
            profile.save()
            target.is_staff = (d["role"] == "administrateur")
            target.save()

            messages.success(request,
                f"✅ Utilisateur '{target.username}' mis à jour.")
            return redirect("liste_utilisateurs")
    else:
        form = ModifierUtilisateurForm(initial={
            "nom_complet": target.get_full_name(),
            "email":       target.email,
            "role":        profile.role,
        })

    return render(request, "core/utilisateurs/form.html", {
        "form":   form,
        "titre":  f"Modifier — {target.username}",
        "mode":   "modification",
        "target": target,
    })


@admin_required
def toggle_utilisateur(request, user_id):
    """Active ou désactive un utilisateur."""
    target = get_object_or_404(User, pk=user_id)

    # Protections
    if target == request.user:
        messages.error(request, "⛔ Vous ne pouvez pas vous désactiver vous-même.")
        return redirect("liste_utilisateurs")
    if target.is_superuser:
        messages.error(request, "⛔ Impossible de désactiver le superadmin.")
        return redirect("liste_utilisateurs")

    target.is_active = not target.is_active
    target.save()

    etat = "activé" if target.is_active else "désactivé"
    messages.success(request, f"✅ Utilisateur '{target.username}' {etat}.")
    return redirect("liste_utilisateurs")


@admin_required
def supprimer_utilisateur(request, user_id):
    target = get_object_or_404(User, pk=user_id)

    if target == request.user:
        messages.error(request, "⛔ Vous ne pouvez pas supprimer votre propre compte.")
        return redirect("liste_utilisateurs")
    if target.is_superuser:
        messages.error(request, "⛔ Impossible de supprimer le superadmin.")
        return redirect("liste_utilisateurs")

    if request.method == "POST":
        username = target.username
        target.delete()
        messages.success(request, f"✅ Utilisateur '{username}' supprimé.")
        return redirect("liste_utilisateurs")

    return render(request, "core/utilisateurs/confirmer_suppression.html", {
        "target": target
    })


# ── Web Push — abonnements ────────────────────────────────────

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse

@login_required
def push_vapid_public_key(request):
    """Retourne la clé publique VAPID pour le client JS."""
    from django.conf import settings
    return JsonResponse({"publicKey": settings.VAPID_PUBLIC_KEY})


@csrf_exempt
@login_required
def push_subscribe(request):
    """Enregistre ou met à jour un abonnement push."""
    if request.method != "POST":
        return JsonResponse({"error": "POST requis"}, status=405)

    try:
        data     = json.loads(request.body)
        endpoint = data.get("endpoint")
        keys     = data.get("keys", {})
        p256dh   = keys.get("p256dh")
        auth_key = keys.get("auth")

        if not all([endpoint, p256dh, auth_key]):
            return JsonResponse({"error": "Données incomplètes"}, status=400)

        from .models import PushSubscription
        sub, created = PushSubscription.objects.update_or_create(
            endpoint = endpoint,
            defaults = {
                "user":       request.user,
                "p256dh":     p256dh,
                "auth":       auth_key,
                "user_agent": request.META.get("HTTP_USER_AGENT", "")[:300],
            }
        )
        action = "créé" if created else "mis à jour"
        return JsonResponse({"status": f"Abonnement {action}"})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
@login_required
def push_unsubscribe(request):
    """Supprime un abonnement push."""
    if request.method != "POST":
        return JsonResponse({"error": "POST requis"}, status=405)

    try:
        from .models import PushSubscription
        data     = json.loads(request.body)
        endpoint = data.get("endpoint")
        PushSubscription.objects.filter(
            user=request.user, endpoint=endpoint
        ).delete()
        return JsonResponse({"status": "Désabonné"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ── Gestion Engins (CRUD) ─────────────────────────────────────

@login_required
def gestion_engins(request):
    engins_actifs   = Engin.objects.filter(actif=True).order_by("id_engin")
    engins_inactifs = Engin.objects.filter(actif=False).order_by("id_engin")

    # Stats par engin
    from django.db.models import Count, Sum as DSum
    stats = {}
    for e in Engin.objects.all():
        stats[e.id_engin] = {
            "nb_ravs":    RavitaillementEngin.objects.filter(engin=e).count(),
            "total_litres": RavitaillementEngin.objects.filter(engin=e).aggregate(
                t=DSum("qte_donnee"))["t"] or 0,
        }

    ctx = {
        "engins_actifs":   engins_actifs,
        "engins_inactifs": engins_inactifs,
        "stats":           stats,
        "total_actifs":    engins_actifs.count(),
        "total_inactifs":  engins_inactifs.count(),
    }
    return render(request, "core/engins/liste.html", ctx)


@admin_required
def creer_engin(request):
    if request.method == "POST":
        from .forms import EnginForm
        form = EnginForm(request.POST)
        if form.is_valid():
            engin = form.save()
            messages.success(request,
                f"✅ Engin '{engin.id_engin}' créé avec succès.")
            return redirect("gestion_engins")
    else:
        from .forms import EnginForm
        form = EnginForm()

    return render(request, "core/engins/form.html", {
        "form":  form,
        "titre": "Ajouter un engin",
        "mode":  "creation",
    })


@admin_required
def modifier_engin(request, id_engin):
    engin = get_object_or_404(Engin, id_engin=id_engin)

    if request.method == "POST":
        from .forms import EnginForm
        form = EnginForm(request.POST, instance=engin, is_edit=True)
        if form.is_valid():
            form.save()
            messages.success(request,
                f"✅ Engin '{engin.id_engin}' mis à jour.")
            return redirect("gestion_engins")
    else:
        from .forms import EnginForm
        form = EnginForm(instance=engin, is_edit=True)

    return render(request, "core/engins/form.html", {
        "form":  form,
        "titre": f"Modifier — {engin.id_engin}",
        "mode":  "modification",
        "engin": engin,
    })


@admin_required
def toggle_engin(request, id_engin):
    engin = get_object_or_404(Engin, id_engin=id_engin)
    engin.actif = not engin.actif
    engin.save()
    etat = "activé" if engin.actif else "désactivé"
    messages.success(request, f"✅ Engin '{engin.id_engin}' {etat}.")
    return redirect("gestion_engins")


@admin_required
def supprimer_engin(request, id_engin):
    engin = get_object_or_404(Engin, id_engin=id_engin)
    nb_ravs = RavitaillementEngin.objects.filter(engin=engin).count()

    if request.method == "POST":
        # Supprimer les sorties stock liées
        OperationStock.objects.filter(
            description__icontains=f"Appro {engin.id_engin}"
        ).delete()
        engin.delete()
        messages.success(request,
            f"✅ Engin '{id_engin}' supprimé ({nb_ravs} ravitaillements supprimés).")
        return redirect("gestion_engins")

    return render(request, "core/engins/confirmer_suppression.html", {
        "engin":   engin,
        "nb_ravs": nb_ravs,
    })


# ── Historique — édition et suppression ───────────────────────

def _recalc_stocks():
    """Recalcule tous les stock_apres après modification."""
    stock = 0
    for op in OperationStock.objects.order_by("date", "cree_le"):
        if op.type == "entree":
            stock += op.quantite
        else:
            stock -= op.quantite
        op.stock_apres = stock
        op.save(update_fields=["stock_apres"])


# ── Opération Stock ───────────────────────────────────────────
@login_required
def edit_operation_stock(request, pk):
    op = get_object_or_404(OperationStock, pk=pk)
    from .forms import EditOperationStockForm

    if request.method == "POST":
        form = EditOperationStockForm(request.POST, instance=op)
        if form.is_valid():
            form.save()
            _recalc_stocks()
            messages.success(request, "✅ Opération stock mise à jour.")
            return redirect("historique")
    else:
        form = EditOperationStockForm(instance=op)

    return render(request, "core/historique/edit_stock.html", {
        "form": form, "op": op
    })


@admin_required
def delete_operation_stock(request, pk):
    op = get_object_or_404(OperationStock, pk=pk)
    if request.method == "POST":
        op.delete()
        _recalc_stocks()
        messages.success(request, "✅ Opération supprimée.")
        return redirect("historique")
    return render(request, "core/historique/confirmer_suppression.html", {
        "objet": f"l'opération stock du {op.date.strftime('%d/%m/%Y')} ({op.quantite:.0f} L)",
        "retour_url": "historique",
    })


# ── Ravitaillement Engin ───────────────────────────────────────
@login_required
def edit_ravitaillement(request, pk):
    rav = get_object_or_404(RavitaillementEngin, pk=pk)
    from .forms import EditRavitaillementEnginForm

    if request.method == "POST":
        form = EditRavitaillementEnginForm(request.POST, instance=rav)
        if form.is_valid():
            r = form.save(commit=False)
            # Recalculer différence index
            r.difference_index = (r.index_actuel or 0) - (r.index_precedent or 0)
            r.save()
            # Mettre à jour la sortie stock correspondante
            OperationStock.objects.filter(
                description__icontains=f"Appro {rav.engin_id}",
                date=rav.date,
            ).update(quantite=r.qte_donnee)
            _recalc_stocks()
            messages.success(request, "✅ Ravitaillement mis à jour.")
            return redirect("historique")
    else:
        form = EditRavitaillementEnginForm(instance=rav)

    return render(request, "core/historique/edit_rav.html", {
        "form": form, "rav": rav
    })


@admin_required
def delete_ravitaillement(request, pk):
    rav = get_object_or_404(RavitaillementEngin, pk=pk)
    if request.method == "POST":
        # Supprimer la sortie stock associée
        OperationStock.objects.filter(
            description__icontains=f"Appro {rav.engin_id}",
            date=rav.date,
            type="sortie",
        ).first() and OperationStock.objects.filter(
            description__icontains=f"Appro {rav.engin_id}",
            date=rav.date,
            type="sortie",
        ).first().delete()
        rav.delete()
        _recalc_stocks()
        messages.success(request, "✅ Ravitaillement supprimé.")
        return redirect("historique")
    return render(request, "core/historique/confirmer_suppression.html", {
        "objet": f"le ravitaillement {rav.engin_id} du {rav.date.strftime('%d/%m/%Y')} ({rav.qte_donnee:.0f} L)",
        "retour_url": "historique",
    })


# ── Consommation Diverse ──────────────────────────────────────
@login_required
def edit_consommation_diverse(request, pk):
    div = get_object_or_404(ConsommationDiverse, pk=pk)
    from .forms import EditConsommationDiverseForm

    if request.method == "POST":
        form = EditConsommationDiverseForm(request.POST, instance=div)
        if form.is_valid():
            form.save()
            OperationStock.objects.filter(
                description__icontains=f"Divers",
                date=div.date,
                type="sortie",
            ).first() and OperationStock.objects.filter(
                description__icontains="Divers",
                date=div.date,
                type="sortie",
            ).update(quantite=form.cleaned_data["quantite"])
            _recalc_stocks()
            messages.success(request, "✅ Consommation diverse mise à jour.")
            return redirect("historique")
    else:
        form = EditConsommationDiverseForm(instance=div)

    return render(request, "core/historique/edit_diverse.html", {
        "form": form, "div": div
    })


@admin_required
def delete_consommation_diverse(request, pk):
    div = get_object_or_404(ConsommationDiverse, pk=pk)
    if request.method == "POST":
        OperationStock.objects.filter(
            description__icontains="Divers",
            date=div.date,
            type="sortie",
        ).first() and OperationStock.objects.filter(
            description__icontains="Divers",
            date=div.date,
            type="sortie",
        ).first().delete()
        div.delete()
        _recalc_stocks()
        messages.success(request, "✅ Consommation diverse supprimée.")
        return redirect("historique")
    return render(request, "core/historique/confirmer_suppression.html", {
        "objet": f"la consommation {div.categorie} du {div.date.strftime('%d/%m/%Y')} ({div.quantite:.0f} L)",
        "retour_url": "historique",
    })


# ── Paramètres ────────────────────────────────────────────────

import base64
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile

@admin_required
def parametres(request):
    from .forms import ParametresForm

    # Charger les valeurs actuelles
    def get_p(cle, defaut=""):
        return Parametre.get(cle, defaut)

    # Chemin logo actuel
    logo_b64 = get_p("logo_base64", "")
    logo_existe = bool(logo_b64)

    if request.method == "POST":
        form = ParametresForm(request.POST, request.FILES)
        if form.is_valid():
            d = form.cleaned_data

            # Sauvegarder les paramètres texte
            for cle, val in [
                ("nom_entreprise",    d["nom_entreprise"]),
                ("unite",             d["unite"]),
                ("seuil_alerte_stock", str(d["seuil_alerte_stock"])),
            ]:
                obj, _ = Parametre.objects.get_or_create(cle=cle)
                obj.valeur = val
                obj.save()

            # Gérer le logo
            if d.get("supprimer_logo"):
                Parametre.objects.filter(cle="logo_base64").delete()
                messages.success(request, "✅ Logo supprimé.")

            elif d.get("logo"):
                fichier = d["logo"]
                # Vérifier taille (max 2MB)
                if fichier.size > 2 * 1024 * 1024:
                    messages.error(request, "❌ Logo trop grand — max 2MB.")
                else:
                    # Stocker en base64 dans Parametre
                    logo_data = base64.b64encode(fichier.read()).decode()
                    ext = fichier.name.rsplit(".", 1)[-1].lower()
                    mime = "image/jpeg" if ext in ("jpg","jpeg") else "image/png"
                    obj, _ = Parametre.objects.get_or_create(cle="logo_base64")
                    obj.valeur = f"data:{mime};base64,{logo_data}"
                    obj.save()
                    # Aussi sauvegarder en fichier pour les PDF
                    _sauvegarder_logo_fichier(fichier)
                    messages.success(request, "✅ Paramètres enregistrés.")
            else:
                messages.success(request, "✅ Paramètres enregistrés.")

            return redirect("parametres")
    else:
        form = ParametresForm(initial={
            "nom_entreprise":    get_p("nom_entreprise", "Mont Gabaon Construction SARLU"),
            "unite":             get_p("unite", "L"),
            "seuil_alerte_stock": float(get_p("seuil_alerte_stock", "500")),
        })

    ctx = {
        "form":        form,
        "logo_existe": logo_existe,
        "logo_b64":    logo_b64,
        "params": {
            "nom_entreprise":    get_p("nom_entreprise", "Mont Gabaon Construction SARLU"),
            "unite":             get_p("unite", "L"),
            "seuil_alerte_stock": get_p("seuil_alerte_stock", "500"),
        }
    }
    return render(request, "core/parametres.html", ctx)


def _sauvegarder_logo_fichier(fichier):
    """Sauvegarde le logo en fichier pour usage dans les PDFs."""
    import sys, os
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.dirname(os.path.dirname(
            os.path.abspath(__file__))))
    assets_dir = os.path.join(base, "assets")
    os.makedirs(assets_dir, exist_ok=True)
    logo_path = os.path.join(assets_dir, "logo_entreprise.png")
    fichier.seek(0)
    with open(logo_path, "wb") as f:
        f.write(fichier.read())


@login_required
def api_parametres(request):
    """API JSON pour récupérer les paramètres publics."""
    from django.http import JsonResponse
    return JsonResponse({
        "nom_entreprise":    Parametre.get("nom_entreprise", "Mont Gabaon Construction SARLU"),
        "seuil_alerte_stock": float(Parametre.get("seuil_alerte_stock", "500")),
        "unite":             Parametre.get("unite", "L"),
    })


# ── Rapports PDF ──────────────────────────────────────────────

@login_required
def rapports(request):
    """Page principale des rapports — sélection mois/année."""
    now    = timezone.localdate()
    annees = range(2024, now.year + 2)
    ctx = {
        "mois":   request.GET.get("mois", now.month),
        "annee":  request.GET.get("annee", now.year),
        "annees": annees,
        "mois_choices": [
            (1,"Janvier"),(2,"Février"),(3,"Mars"),(4,"Avril"),
            (5,"Mai"),(6,"Juin"),(7,"Juillet"),(8,"Août"),
            (9,"Septembre"),(10,"Octobre"),(11,"Novembre"),(12,"Décembre"),
        ],
    }
    return render(request, "core/rapports.html", ctx)


def _pdf_response(buf, filename):
    """Retourne une HttpResponse PDF téléchargeable."""
    response = HttpResponse(buf.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def pdf_rapport_entrees(request):
    mois  = request.GET.get("mois",  timezone.localdate().month)
    annee = request.GET.get("annee", timezone.localdate().year)
    try:
        from .pdf_service import generer_rapport_entrees
        buf = generer_rapport_entrees(mois, annee)
        return _pdf_response(buf, f"rapport_entrees_{annee}_{int(mois):02d}.pdf")
    except Exception as e:
        messages.error(request, f"❌ Erreur génération PDF : {e}")
        return redirect("rapports")


@login_required
def pdf_attestation(request):
    mois       = request.GET.get("mois",        timezone.localdate().month)
    annee      = request.GET.get("annee",       timezone.localdate().year)
    responsable = request.GET.get("responsable", "")
    try:
        from .pdf_service import generer_attestation
        buf = generer_attestation(mois, annee, responsable)
        return _pdf_response(buf, f"attestation_{annee}_{int(mois):02d}.pdf")
    except Exception as e:
        messages.error(request, f"❌ Erreur génération PDF : {e}")
        return redirect("rapports")


@login_required
def pdf_rapport_mensuel(request):
    mois  = request.GET.get("mois",  timezone.localdate().month)
    annee = request.GET.get("annee", timezone.localdate().year)
    try:
        from .pdf_service import generer_rapport_mensuel
        buf = generer_rapport_mensuel(mois, annee)
        return _pdf_response(buf, f"rapport_mensuel_{annee}_{int(mois):02d}.pdf")
    except Exception as e:
        messages.error(request, f"❌ Erreur génération PDF : {e}")
        return redirect("rapports")
